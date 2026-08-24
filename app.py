# -*- coding: utf-8 -*-
# SUBMISSION_TEMPLATE_ALL_SHEETS_V5 = 2026-08-14
# MergedCell-safe version

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import datetime
import base64
import glob
from io import BytesIO

import matplotlib.pyplot as plt
from matplotlib import font_manager
from pathlib import Path
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from data.loader import load_data, load_media_master
from logic.forecast import forecast_cv
from logic.simulation import simulate_plan
from logic.optimize import optimize_budget
from logic.analytics import prepare_monthly_performance, prepare_unit_price_band_matrix


@st.cache_data(show_spinner=False, max_entries=3)
def _load_history_cached(file_bytes: bytes, exclude_compensation: bool) -> pd.DataFrame:
    """アップロードCSVの再読込を防ぎ、Streamlit再実行時のメモリ急増を抑える。"""
    return load_data(BytesIO(file_bytes), exclude_compensation=exclude_compensation)


@st.cache_data(show_spinner=False, max_entries=3)
def _load_master_cached(file_bytes: bytes):
    """CPNマスタ/媒体名マスタを同じExcelから一度だけ読み込む。"""
    buf = BytesIO(file_bytes)
    cpn = pd.read_excel(buf, sheet_name="CPNマスタ", engine="openpyxl")
    buf.seek(0)
    media = load_media_master(buf, sheet_name="媒体名マスタ")
    return cpn, media


# -----------------------
# ✅ 日付フォーマット
# -----------------------
def format_date(df, col="date"):
    return (
        pd.to_datetime(df[col])
        .dt.strftime("%Y/%m/%d")
        .str.replace("/0", "/", regex=False)
    )


# -----------------------
# ✅ 帳票形式
# -----------------------
def create_report_table(df):
    """
    松竹梅・最適プラン表示用の高速版。
    媒体×プランごとのループとDataFrame大量生成をやめ、
    melt + pivot_table で一括変換する。
    """
    if df.empty:
        return pd.DataFrame()

    work = df[["date", "media", "plan", "cv", "cost", "cpa"]].copy()

    long_df = work.melt(
        id_vars=["media", "plan", "date"],
        value_vars=["cv", "cost", "cpa"],
        var_name="metric",
        value_name="value",
    )

    metric_map = {
        "cv": "CV",
        "cost": "COST",
        "cpa": "CPA",
    }
    metric_order = {
        "CV": 0,
        "COST": 1,
        "CPA": 2,
    }

    long_df["metric"] = long_df["metric"].map(metric_map)
    long_df["_metric_order"] = long_df["metric"].map(metric_order)

    result = (
        long_df.pivot_table(
            index=["media", "plan", "metric", "_metric_order"],
            columns="date",
            values="value",
            aggfunc="sum",
        )
        .reset_index()
        .sort_values(
            ["media", "plan", "_metric_order"],
            kind="stable",
        )
        .drop(columns="_metric_order")
        .reset_index(drop=True)
    )

    date_cols = [
        c for c in result.columns
        if c not in ["media", "plan", "metric"]
    ]
    result = result[["media", "plan", "metric"] + date_cols]

    result["media"] = result["media"].mask(
        result["media"].duplicated()
    )

    # planは3行（CV/COST/CPA）の先頭だけ表示
    result["plan"] = result["plan"].mask(
        result["plan"].eq(result["plan"].shift())
        & result["media"].isna()
    )

    return result


# -----------------------
# ✅ Excel
# -----------------------
def to_excel_multi(sim_df, opt_df):

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        create_report_table(sim_df).to_excel(writer, sheet_name="松竹梅", index=False)
        create_report_table(opt_df).to_excel(writer, sheet_name="最適", index=False)

    return output.getvalue()


# -----------------------
# ✅ 提出用Excel（最適プランのみ / 添付テンプレ全シート再現）
# -----------------------
def _template_path() -> Path:
    base = Path(__file__).resolve().parent
    candidates = [
        base / "assets" / "submission_template_v2.xlsx",
        base / "submission_template_v2.xlsx",
        base / "assets" / "submission_template_fast.xlsx",
        base / "submission_template_fast.xlsx",
        base / "assets" / "submission_template.xlsx",
        base / "submission_template.xlsx",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError(
        "提出用テンプレートが見つかりません。assets/submission_template.xlsx を配置してください。"
    )


def _safe_number(value, default=0.0):
    value = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return default if pd.isna(value) else float(value)


def _set_date_slots(ws, row, first_col, slot_count, dates, total_col=None):
    """テンプレの日付セル書式を維持しつつ、予測期間の日付へ差し替える。"""
    for i in range(slot_count):
        cell = ws.cell(row, first_col + i)

        # 結合セルの左上以外には書き込まない
        if isinstance(cell, MergedCell):
            continue

        if i < len(dates):
            cell.value = dates[i].to_pydatetime()
            cell.number_format = "m/d"
        else:
            cell.value = None

    if total_col:
        cell = ws.cell(row, total_col)
        if not isinstance(cell, MergedCell):
            cell.value = "Total"


def _set_value(ws, row, col, value):
    """
    結合セル対策付きの値セット。
    左上セル以外の MergedCell に当たった場合は何もしない。
    """
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _set_percent(ws, row, col, value):
    """0.583 をExcel上で58.3%表示にする。"""
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    cell.number_format = "0.0%"




def _calculate_media_approval_rates(history_df: pd.DataFrame) -> dict:
    """
    過去実績の『成果承認フラグ = Y』を発行数として、媒体別承認率を算出する。

    承認率 = 成果承認フラグYの件数合計 / 全発生件数合計

    媒体に承認実績がない場合は全媒体の加重承認率をフォールバックに使う。
    全体でも承認実績がない場合はエラーにする。
    """
    required = {"media", "approved_cv", "approval_base_cv"}
    missing = required - set(history_df.columns)

    if missing:
        raise ValueError(
            "承認率算出用データがありません。"
            "data/loader.py を承認率対応版へ更新してください。"
        )

    work = history_df[
        ["media", "approved_cv", "approval_base_cv"]
    ].copy()

    work["approved_cv"] = pd.to_numeric(
        work["approved_cv"],
        errors="coerce",
    )
    work["approval_base_cv"] = pd.to_numeric(
        work["approval_base_cv"],
        errors="coerce",
    )

    valid = work[
        work["approval_base_cv"].notna()
        & (work["approval_base_cv"] > 0)
    ].copy()

    if valid.empty:
        source_col = ""
        if "approval_source_column" in history_df.columns:
            sources = (
                history_df["approval_source_column"]
                .dropna()
                .astype(str)
                .str.strip()
            )
            sources = sources[sources != ""]
            if not sources.empty:
                source_col = sources.iloc[0]

        if source_col:
            raise ValueError(
                f"承認列『{source_col}』は見つかりましたが、"
                "Y/N・1/0・承認/否認として判定できる実績がありません。"
            )

        raise ValueError(
            "実績CSVから承認判定列を見つけられませんでした。"
            "『承認フラグ』『承認状況』『承認ステータス』『承認』"
            "などの列を確認してください。"
        )

    total_base = valid["approval_base_cv"].sum()
    total_approved = valid["approved_cv"].fillna(0).sum()

    overall_rate = (
        float(total_approved / total_base)
        if total_base > 0
        else 0.0
    )
    overall_rate = min(max(overall_rate, 0.0), 1.0)

    media_rates = (
        valid.groupby("media", as_index=False)
        .agg(
            approved_cv=("approved_cv", "sum"),
            approval_base_cv=("approval_base_cv", "sum"),
        )
    )

    media_rates["approval_rate"] = (
        media_rates["approved_cv"]
        / media_rates["approval_base_cv"]
    ).clip(0, 1)

    rate_map = dict(
        zip(
            media_rates["media"].astype(str),
            media_rates["approval_rate"].astype(float),
        )
    )

    # どの媒体でも必ず率が取れるよう全体率を保持
    rate_map["__overall__"] = overall_rate
    return rate_map



def _calculate_period_media_metrics(history_df: pd.DataFrame):
    """
    媒体別に定常・マジ得の過去実績指標を作る。

    定常承認率:
        定常/通常期間の 成果承認フラグY件数 / 全件数

    マジ得承認率:
        マジ得期間の 成果承認フラグY件数 / 全件数

    マジ得単価:
        マジ得期間の cost / cv

    媒体に該当期間実績がない場合は、その期間の全媒体加重値をフォールバック。
    """
    required = {
        "media", "CPN名", "cv", "cost",
        "approved_cv", "approval_base_cv",
    }
    missing = required - set(history_df.columns)
    if missing:
        raise ValueError(
            "定常・マジ得指標の算出に必要な列がありません: "
            + ", ".join(sorted(missing))
        )

    work = history_df[
        ["media", "CPN名", "cv", "cost", "approved_cv", "approval_base_cv"]
    ].copy()

    work["media"] = work["media"].astype(str)
    work["CPN名"] = work["CPN名"].astype(str).str.strip()
    for col in ["cv", "cost", "approved_cv", "approval_base_cv"]:
        work[col] = pd.to_numeric(work[col], errors="coerce").fillna(0)

    def build(mask):
        sub = work[mask].copy()

        if sub.empty:
            return {}, 0.0, {}, 0.0

        agg = (
            sub.groupby("media", as_index=False)
            .agg(
                cv=("cv", "sum"),
                cost=("cost", "sum"),
                approved_cv=("approved_cv", "sum"),
                approval_base_cv=("approval_base_cv", "sum"),
            )
        )

        agg["approval_rate"] = (
            agg["approved_cv"]
            / agg["approval_base_cv"].replace(0, pd.NA)
        ).fillna(0).clip(0, 1)

        agg["unit_price"] = (
            agg["cost"]
            / agg["cv"].replace(0, pd.NA)
        ).fillna(0)

        total_base = sub["approval_base_cv"].sum()
        overall_rate = (
            float(sub["approved_cv"].sum() / total_base)
            if total_base > 0 else 0.0
        )
        total_cv = sub["cv"].sum()
        overall_unit = (
            float(sub["cost"].sum() / total_cv)
            if total_cv > 0 else 0.0
        )

        rate_map = dict(zip(agg["media"], agg["approval_rate"].astype(float)))
        unit_map = dict(zip(agg["media"], agg["unit_price"].astype(float)))
        return rate_map, overall_rate, unit_map, overall_unit

    normal_mask = work["CPN名"].isin(["通常", "定常"])
    magi_mask = work["CPN名"].eq("マジ得")

    normal_rate_map, normal_rate_all, normal_unit_map, normal_unit_all = build(normal_mask)
    magi_rate_map, magi_rate_all, magi_unit_map, magi_unit_all = build(magi_mask)

    return {
        "normal_rate": normal_rate_map,
        "normal_rate_all": normal_rate_all,
        "normal_unit": normal_unit_map,
        "normal_unit_all": normal_unit_all,
        "magi_rate": magi_rate_map,
        "magi_rate_all": magi_rate_all,
        "magi_unit": magi_unit_map,
        "magi_unit_all": magi_unit_all,
    }


def _future_period_name(date_value, future_cpn_map, selected_cpn):
    """
    未来日を定常/マジ得に分類する。
    CPNマスタに当日登録があればそれを優先。
    未登録の場合はUI選択CPNを使用。
    """
    cpn_name = future_cpn_map.get(
        pd.Timestamp(date_value).normalize(),
        selected_cpn,
    )
    cpn_name = str(cpn_name).strip()
    return "マジ得" if cpn_name == "マジ得" else "定常"



def _build_manual_settings_defaults(
    opt_summary: pd.DataFrame,
    history_df: pd.DataFrame,
    selected_cpn: str,
) -> pd.DataFrame:
    """
    最適プラン / 提案用Excelの計算値を、手動設定テーブルの初期値へ変換する。

    編集可能:
      今回プラン採用グロス単価
      今回プラン採用承認率
      今回採用件数
      費用

    自動計算:
      承認件数 = 今回採用件数 × 今回プラン採用承認率
      発行CPA   = 費用 ÷ 承認件数
    """
    plan = opt_summary.copy()
    plan["media"] = plan["media"].astype(str)
    plan["cv"] = pd.to_numeric(plan["cv"], errors="coerce").fillna(0)
    plan["cost"] = pd.to_numeric(plan["cost"], errors="coerce").fillna(0)

    totals = (
        plan.groupby("media", as_index=False)
        .agg(
            plan_cv=("cv", "sum"),
            plan_cost=("cost", "sum"),
        )
    )
    totals["opt_unit"] = (
        totals["plan_cost"]
        / totals["plan_cv"].replace(0, pd.NA)
    ).fillna(0)

    period = _calculate_period_media_metrics(history_df)

    # SID
    sid_map = {}
    if "SID" in history_df.columns:
        sid_source = history_df[["media", "SID"]].copy()
        sid_source["media"] = sid_source["media"].astype(str)
        sid_source["SID"] = (
            sid_source["SID"]
            .fillna("")
            .astype(str)
            .str.strip()
        )
        sid_source = sid_source[sid_source["SID"] != ""]
        sid_map = (
            sid_source.groupby("media")["SID"]
            .agg(lambda x: " / ".join(dict.fromkeys(x.tolist())))
            .to_dict()
        )

    rows = []
    for r in totals.itertuples():
        media = str(r.media)
        opt_unit = float(r.opt_unit or 0)

        # 複合施策時は、施策別予測CVを重みにして承認率・グロス単価をブレンドする。
        media_plan = plan[plan["media"].eq(media)].copy()
        if "CPN名" in media_plan.columns and not media_plan.empty:
            mix = media_plan.groupby("CPN名", as_index=False)["cv"].sum()
            weighted_rate = 0.0
            weighted_gross = 0.0
            weight_total = float(mix["cv"].sum())
            for mix_row in mix.itertuples():
                cpn_name = str(mix_row.CPN名).strip()
                weight = float(mix_row.cv or 0)
                if cpn_name == "マジ得":
                    cpn_rate = period["magi_rate"].get(media, period["magi_rate_all"])
                    cpn_gross = period["magi_unit"].get(media, period["magi_unit_all"])
                    if cpn_gross <= 0:
                        cpn_gross = opt_unit
                else:
                    cpn_rate = period["normal_rate"].get(media, period["normal_rate_all"])
                    cpn_gross = opt_unit
                weighted_rate += weight * float(cpn_rate or 0)
                weighted_gross += weight * float(cpn_gross or 0)
            rate = weighted_rate / weight_total if weight_total else 0.0
            gross_unit = weighted_gross / weight_total if weight_total else opt_unit
        elif selected_cpn == "マジ得":
            rate = period["magi_rate"].get(media, period["magi_rate_all"])
            gross_unit = period["magi_unit"].get(media, period["magi_unit_all"])
            if gross_unit <= 0:
                gross_unit = opt_unit
        else:
            rate = period["normal_rate"].get(media, period["normal_rate_all"])
            gross_unit = opt_unit

        adopted_count = float(r.plan_cv or 0)
        approved_count = adopted_count * float(rate or 0)
        cost = approved_count * float(gross_unit or 0)
        issue_cpa = cost / approved_count if approved_count else 0

        rows.append(
            {
                "SID": sid_map.get(media, ""),
                "媒体名": media,
                "今回プラン採用グロス単価": round(gross_unit),
                "今回プラン採用承認率": float(rate or 0),
                "今回採用件数": round(adopted_count),
                "費用": round(cost),
                "承認件数": round(approved_count, 1),
                "発行CPA": round(issue_cpa),
            }
        )

    result = pd.DataFrame(rows)

    if not result.empty:
        result = result.sort_values(
            ["今回採用件数", "費用", "媒体名"],
            ascending=[False, False, True],
        ).reset_index(drop=True)

    return result


def _normalize_manual_settings(df: pd.DataFrame) -> pd.DataFrame:
    """手動設定の入力値を安全に数値化し、派生値を再計算する。"""
    out = df.copy()

    required_cols = [
        "SID",
        "媒体名",
        "今回プラン採用グロス単価",
        "今回プラン採用承認率",
        "今回採用件数",
        "費用",
    ]
    for col in required_cols:
        if col not in out.columns:
            out[col] = "" if col in {"SID", "媒体名"} else 0

    for col in [
        "今回プラン採用グロス単価",
        "今回プラン採用承認率",
        "今回採用件数",
        "費用",
    ]:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)

    out["今回プラン採用グロス単価"] = (
        out["今回プラン採用グロス単価"].clip(lower=0)
    )
    out["今回プラン採用承認率"] = (
        out["今回プラン採用承認率"].clip(lower=0, upper=1)
    )
    out["今回採用件数"] = out["今回採用件数"].clip(lower=0)
    out["費用"] = out["費用"].clip(lower=0)

    out["承認件数"] = (
        out["今回採用件数"]
        * out["今回プラン採用承認率"]
    )

    issue_denominator = out["承認件数"].where(out["承認件数"].ne(0))
    out["発行CPA"] = (
        out["費用"] / issue_denominator
    ).fillna(0)

    out["今回採用件数"] = out["今回採用件数"].round(0)
    out["費用"] = out["費用"].round(0)
    out["承認件数"] = out["承認件数"].round(1)
    out["発行CPA"] = out["発行CPA"].round(0)

    return out[
        [
            "SID",
            "媒体名",
            "今回プラン採用グロス単価",
            "今回プラン採用承認率",
            "今回採用件数",
            "費用",
            "承認件数",
            "発行CPA",
        ]
    ]


def _manual_settings_signature(df: pd.DataFrame):
    """session_state更新判定用。"""
    cols = [
        "媒体名",
        "今回プラン採用グロス単価",
        "今回プラン採用承認率",
        "今回採用件数",
        "費用",
    ]
    if df is None or df.empty:
        return ()
    work = df[cols].copy()
    return tuple(
        tuple(row)
        for row in work.astype(object).itertuples(index=False, name=None)
    )


def render_manual_settings(
    opt_summary,
    history_df,
    selected_cpn,
    calc_key,
):
    """
    手動設定エディタの安全版。

    st.fragment / st.rerun は使用しない。
    data_editor の編集時はStreamlit標準の再実行に任せる。
    予測・最適化結果は既存のsession_stateキャッシュを再利用するため、
    重い再計算は発生しない。
    """
    if st.session_state.get("_manual_calc_key") != calc_key:
        st.session_state["_manual_calc_key"] = calc_key
        st.session_state["_manual_settings"] = _build_manual_settings_defaults(
            opt_summary=opt_summary,
            history_df=history_df,
            selected_cpn=selected_cpn,
        )

        # 計算条件が変わった時だけEditorのwidget stateもリセット
        old_widget_key = st.session_state.get("_manual_widget_key")
        if old_widget_key:
            st.session_state.pop(old_widget_key, None)

        st.session_state["_manual_widget_key"] = (
            f"manual_settings_editor_{abs(hash(str(calc_key)))}"
        )

    current = _normalize_manual_settings(
        st.session_state.get(
            "_manual_settings",
            _build_manual_settings_defaults(
                opt_summary=opt_summary,
                history_df=history_df,
                selected_cpn=selected_cpn,
            ),
        )
    )

    widget_key = st.session_state.get(
        "_manual_widget_key",
        f"manual_settings_editor_{abs(hash(str(calc_key)))}",
    )

    edited = st.data_editor(
        current,
        key=widget_key,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        disabled=[
            "SID",
            "媒体名",
            "承認件数",
            "発行CPA",
        ],
        column_config={
            "SID": st.column_config.TextColumn(
                "SID",
                width="small",
            ),
            "媒体名": st.column_config.TextColumn(
                "媒体名",
                width="large",
            ),
            "今回プラン採用グロス単価": st.column_config.NumberColumn(
                "今回プラン採用グロス単価",
                min_value=0,
                step=100,
                format="¥%d",
            ),
            "今回プラン採用承認率": st.column_config.NumberColumn(
                "今回プラン採用承認率",
                min_value=0.0,
                max_value=1.0,
                step=0.001,
                format="%.3f",
                help="0.50 = 50% として入力",
            ),
            "今回採用件数": st.column_config.NumberColumn(
                "今回採用件数",
                min_value=0,
                step=1,
                format="%d",
            ),
            "費用": st.column_config.NumberColumn(
                "費用",
                min_value=0,
                step=1000,
                format="¥%d",
            ),
            "承認件数": st.column_config.NumberColumn(
                "承認件数",
                format="%.1f",
            ),
            "発行CPA": st.column_config.NumberColumn(
                "発行CPA",
                format="¥%d",
            ),
        },
    )

    normalized = _normalize_manual_settings(edited)
    st.session_state["_manual_settings"] = normalized

    # 派生値はEditorの下に必ず最新値を表示。
    total_count = normalized["今回採用件数"].sum()
    total_issue = normalized["承認件数"].sum()
    total_cost = normalized["費用"].sum()
    overall_rate = total_issue / total_count if total_count else 0
    overall_cpa = total_cost / total_issue if total_issue else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("今回採用件数 合計", f"{total_count:,.0f}")
    c2.metric("承認件数 合計", f"{total_issue:,.1f}")
    c3.metric("全体承認率", f"{overall_rate:.1%}")
    c4.metric("全体発行CPA", f"¥{overall_cpa:,.0f}")

    # 自動計算結果の確認用
    derived = normalized[
        [
            "媒体名",
            "承認件数",
            "発行CPA",
        ]
    ].copy()

    with st.expander("自動計算結果を確認"):
        st.dataframe(
            derived,
            width="stretch",
            hide_index=True,
            column_config={
                "媒体名": st.column_config.TextColumn("媒体名"),
                "承認件数": st.column_config.NumberColumn(
                    "承認件数",
                    format="%.1f",
                ),
                "発行CPA": st.column_config.NumberColumn(
                    "発行CPA",
                    format="¥%d",
                ),
            },
        )


def create_submission_excel(opt_summary, history_df, cpn_master, manual_settings, start_date, end_date, selected_cpn, opt_mode):
    """
    添付された提出用Excelそのものをテンプレートとして使い、
    最適プランを初期値とした手動設定の結果を提出用Excelへ反映する。

    - SID: 実績データF列を loader.py で保持した history_df["SID"]
    - 件数: 最適プランCV
    - コスト: 最適プランcost
    - 承認率: 過去実績から媒体別に算出
    - 発行: 最適プランForecast × 媒体別承認率
    - Actual: 提出時点では未入力

    ※テンプレ内の既存媒体名や数値は、出力時に対象範囲をクリアしてから
      最適プランの値へ置き換える。
    ※元テンプレートそのものは変更しない。
    """
    plan = opt_summary.copy()
    plan["date"] = pd.to_datetime(plan["date"], errors="coerce").dt.normalize()
    plan = plan.dropna(subset=["date", "media"])
    plan["media"] = plan["media"].astype(str)
    plan["cv"] = pd.to_numeric(plan["cv"], errors="coerce").fillna(0)
    plan["cost"] = pd.to_numeric(plan["cost"], errors="coerce").fillna(0)

    # 複合施策では「開始～終了の連続日」ではなく、実際に予測した対象日だけを出力する。
    dates = sorted(pd.Timestamp(x).normalize() for x in plan["date"].dropna().unique())
    if len(dates) > 33:
        raise ValueError("提出用テンプレートは最大33日分です。施策①+施策②の対象日数を33日以内にしてください。")

    daily = (
        plan.groupby(["date", "media"], as_index=False)
        .agg(cv=("cv", "sum"), cost=("cost", "sum"))
    )
    daily["cpa"] = (
        (daily["cost"] / daily["cv"])
        .replace([float("inf"), float("-inf")], 0)
        .fillna(0)
    )

    base_media_totals = (
        daily.groupby("media", as_index=False)
        .agg(
            total_cv=("cv", "sum"),
            total_cost=("cost", "sum"),
        )
    )
    base_media_totals["media"] = base_media_totals["media"].astype(str)

    # ---------------------------------------------------------
    # 手動設定を最終提出値として採用。
    # 日次件数は、元の最適プラン日次構成比を維持して再配分する。
    # ---------------------------------------------------------
    manual = _normalize_manual_settings(manual_settings)
    manual["媒体名"] = manual["媒体名"].astype(str)

    base_media_set = set(base_media_totals["media"].tolist())
    manual = manual[manual["媒体名"].isin(base_media_set)].copy()

    manual = manual.sort_values(
        ["今回採用件数", "費用", "媒体名"],
        ascending=[False, False, True],
    ).reset_index(drop=True)

    media_list = manual["媒体名"].tolist()

    MAX_TEMPLATE_MEDIA = 150
    if len(media_list) > MAX_TEMPLATE_MEDIA:
        raise ValueError(
            f"提出用Excelは最大{MAX_TEMPLATE_MEDIA}媒体まで対応しています。"
            f"現在は{len(media_list)}媒体です。"
        )

    manual_count_map = dict(
        zip(
            manual["媒体名"],
            manual["今回採用件数"].astype(float),
        )
    )
    manual_rate_map = dict(
        zip(
            manual["媒体名"],
            manual["今回プラン採用承認率"].astype(float),
        )
    )
    manual_cost_total_map = dict(
        zip(
            manual["媒体名"],
            manual["費用"].astype(float),
        )
    )
    manual_gross_unit_map = dict(
        zip(
            manual["媒体名"],
            manual["今回プラン採用グロス単価"].astype(float),
        )
    )

    original_total_cv_map = dict(
        zip(
            base_media_totals["media"],
            base_media_totals["total_cv"].astype(float),
        )
    )

    # 元日次Forecastを今回採用件数へ比例配分。
    cv_map = {}
    for row in daily.itertuples():
        dt = pd.Timestamp(row.date).normalize()
        media = str(row.media)
        original_total = original_total_cv_map.get(media, 0.0)
        adopted_total = manual_count_map.get(media, 0.0)

        scale = (
            adopted_total / original_total
            if original_total > 0
            else 0.0
        )
        cv_map[(dt, media)] = float(row.cv) * scale

    # 端数差が出ても媒体Totalが手動入力値と一致するよう、最後の日へ差分を寄せる。
    dates_by_media = {}
    for dt, media in cv_map.keys():
        dates_by_media.setdefault(media, []).append(dt)

    for media in media_list:
        media_dates = sorted(dates_by_media.get(media, []))
        if not media_dates:
            continue

        current_total = sum(
            cv_map.get((dt, media), 0.0)
            for dt in media_dates
        )
        diff = manual_count_map.get(media, 0.0) - current_total
        cv_map[(media_dates[-1], media)] = (
            cv_map.get((media_dates[-1], media), 0.0)
            + diff
        )

    # 過去実績の定常・マジ得指標は、提出用Excelの参考列用に残す。
    period_metrics = _calculate_period_media_metrics(history_df)
    normal_rate_map = period_metrics["normal_rate"]
    magi_rate_map = period_metrics["magi_rate"]
    normal_rate_all = period_metrics["normal_rate_all"]
    magi_rate_all = period_metrics["magi_rate_all"]
    magi_unit_map = period_metrics["magi_unit"]
    magi_unit_all = period_metrics["magi_unit_all"]

    # 未来日ごとのCPN区分。複合施策時はUI選択を保持したplan側を最優先。
    future_cpn_map = {}
    if "CPN名" in plan.columns:
        fm = plan[["date", "CPN名"]].dropna(subset=["date"]).drop_duplicates(subset=["date"], keep="last")
        future_cpn_map = dict(zip(fm["date"], fm["CPN名"]))
    elif cpn_master is not None and not cpn_master.empty:
        fm = cpn_master[["日付", "CPN名"]].copy()
        fm["日付"] = pd.to_datetime(fm["日付"], errors="coerce").dt.normalize()
        fm = fm.dropna(subset=["日付"])
        future_cpn_map = dict(zip(fm["日付"], fm["CPN名"]))

    # 発行数 = 手動設定後Forecast × 手動承認率
    issue_map = {}
    for (dt, media), cv in cv_map.items():
        rate = manual_rate_map.get(media, 0.0)
        issue_map[(dt, media)] = cv * rate

    # 費用はユーザー入力総額を正として、日次承認件数の構成比で配分。
    expected_cost_map = {}
    for media in media_list:
        media_dates = sorted(dates_by_media.get(media, []))
        total_cost = manual_cost_total_map.get(media, 0.0)
        total_issue = sum(
            issue_map.get((dt, media), 0.0)
            for dt in media_dates
        )

        if total_issue > 0:
            for dt in media_dates:
                share = issue_map.get((dt, media), 0.0) / total_issue
                expected_cost_map[(dt, media)] = total_cost * share
        else:
            total_cv = sum(
                cv_map.get((dt, media), 0.0)
                for dt in media_dates
            )
            for dt in media_dates:
                share = (
                    cv_map.get((dt, media), 0.0) / total_cv
                    if total_cv > 0
                    else 0.0
                )
                expected_cost_map[(dt, media)] = total_cost * share

        # 費用Totalの丸め差は最後の日へ寄せる。
        if media_dates:
            assigned = sum(
                expected_cost_map.get((dt, media), 0.0)
                for dt in media_dates
            )
            diff = total_cost - assigned
            last_dt = media_dates[-1]
            expected_cost_map[(last_dt, media)] = (
                expected_cost_map.get((last_dt, media), 0.0)
                + diff
            )

    # Promotion Detail等は今回採用グロス単価を表示。
    period_unit_map = {
        (dt, media): manual_gross_unit_map.get(media, 0.0)
        for (dt, media) in cv_map.keys()
    }

    total_cv_by_date = {}
    total_issue_by_date = {}
    total_cost_by_date = {}

    for (dt, media), cv in cv_map.items():
        total_cv_by_date[dt] = total_cv_by_date.get(dt, 0.0) + cv
        total_issue_by_date[dt] = (
            total_issue_by_date.get(dt, 0.0)
            + issue_map.get((dt, media), 0.0)
        )
        total_cost_by_date[dt] = (
            total_cost_by_date.get(dt, 0.0)
            + expected_cost_map.get((dt, media), 0.0)
        )

    # 参考列用：今回採用件数を未来CPN区分に応じて分割。
    normal_forecast_by_media = {m: 0.0 for m in media_list}
    magi_forecast_by_media = {m: 0.0 for m in media_list}

    for (dt, media), cv in cv_map.items():
        period = _future_period_name(
            dt,
            future_cpn_map,
            selected_cpn,
        )
        if period == "マジ得":
            magi_forecast_by_media[media] += cv
        else:
            normal_forecast_by_media[media] += cv

    normal_issue_by_media = {
        m: (
            normal_forecast_by_media[m]
            * manual_rate_map.get(m, 0.0)
        )
        for m in media_list
    }
    magi_issue_by_media = {
        m: (
            magi_forecast_by_media[m]
            * manual_rate_map.get(m, 0.0)
        )
        for m in media_list
    }

    # 後段互換用。
    opt_unit_map = manual_gross_unit_map
    cpa_map = {
        key: (
            expected_cost_map.get(key, 0.0)
            / issue_map.get(key, 0.0)
            if issue_map.get(key, 0.0) > 0
            else 0.0
        )
        for key in cv_map.keys()
    }

    # SIDは実績CSVのF列を正とする。
    sid_map = {}
    if "SID" in history_df.columns:
        sid_source = history_df[["media", "SID"]].copy()
        sid_source["media"] = sid_source["media"].astype(str)
        sid_source["SID"] = sid_source["SID"].fillna("").astype(str).str.strip()
        sid_source = sid_source[sid_source["SID"] != ""]
        sid_map = (
            sid_source.groupby("media")["SID"]
            .agg(lambda x: " / ".join(dict.fromkeys(x.tolist())))
            .to_dict()
        )

    template = _template_path()
    wb_out = load_workbook(template)

    # 今回は短縮承認系2シートを成果物から除外。
    for remove_name in ["短縮承認除外", "短縮承認日程"]:
        if remove_name in wb_out.sheetnames:
            del wb_out[remove_name]

    # 高速テンプレートは数式依存を最小化しているため、
    # 保存時の強制フル再計算指定は行わない。

    main_old_name = "10月（既存移管合算）"
    main_ws = (
        wb_out[main_old_name]
        if main_old_name in wb_out.sheetnames
        else wb_out.worksheets[2]
    )

    # テンプレにあるSID→媒体区分の対応を先に取得してから、
    # 明細の既存値をクリアする。
    template_type_by_sid = {}
    template_type_by_media = {}
    for r in range(1, main_ws.max_row + 1):
        sid = main_ws.cell(r, 3).value
        media = main_ws.cell(r, 4).value
        media_type = main_ws.cell(r, 19).value
        if sid not in (None, "") and media_type not in (None, ""):
            template_type_by_sid[str(sid).strip()] = str(media_type).strip()
        if media not in (None, "") and media_type not in (None, ""):
            template_type_by_media[str(media).strip()] = str(media_type).strip()

    media_type_map = {}
    for media in media_list:
        sid = sid_map.get(media, "").split(" / ")[0].strip()
        media_type_map[media] = (
            template_type_by_sid.get(sid)
            or template_type_by_media.get(media)
            or "ポイントサイト"
        )

    # =========================================================
    # 1) メインシート：○月（既存移管合算）
    # =========================================================
    new_main_name = f"{pd.Timestamp(start_date).month}月（既存移管合算）"
    main_ws.title = new_main_name

    # 最小テンプレートは他シートから旧シート名を参照する数式を持たないため、
    # 全ワークシート・全セルの走査は行わない。
    # これにより提出用Excel生成時の処理時間を大幅に削減する。

    first_date_col = 25  # Y
    total_col = 58       # BF
    date_slots = 33

    # 上部サマリの日付・Target/Actual/GAP
    _set_date_slots(main_ws, 2, first_date_col, date_slots, dates, total_col)
    _set_date_slots(main_ws, 7, first_date_col, date_slots, dates, total_col)

    for i in range(date_slots):
        col = first_date_col + i
        if i < len(dates):
            dt = dates[i]
            target = round(total_cv_by_date.get(dt.normalize(), 0))
            _set_value(main_ws, 3, col, target)
            _set_value(main_ws, 4, col, None)
            _set_value(main_ws, 5, col, -target)
            _set_value(main_ws, 8, col, "月火水木金土日"[dt.weekday()])
        else:
            for rr in (3, 4, 5, 8):
                _set_value(main_ws, rr, col, None)

    _set_value(main_ws, 3, total_col, round(sum(total_cv_by_date.values())))
    _set_value(main_ws, 4, total_col, 0)
    _set_value(main_ws, 5, total_col, -round(sum(total_cv_by_date.values())))

    # 媒体区分ごとの上部4ブロックを最適プランから再集計。
    group_rows = {}
    for r in range(8, 21):
        label = main_ws.cell(r, 4).value
        metric = main_ws.cell(r, 24).value
        if (
            isinstance(label, str)
            and "合計" in label
            and metric == "Daily Target (Initiative)"
        ):
            key = label.replace("【", "").replace("】合計", "").strip()
            group_rows[key] = r

    for group_name, start_row in group_rows.items():
        members = [m for m in media_list if media_type_map.get(m) == group_name]

        for i in range(date_slots):
            col = first_date_col + i
            if i < len(dates):
                dt = dates[i].normalize()
                val = round(sum(cv_map.get((dt, m), 0) for m in members))
                _set_value(main_ws, start_row, col, val)
                _set_value(main_ws, start_row + 1, col, None)
                _set_value(main_ws, start_row + 2, col, -val)
            else:
                for rr in range(
                    start_row,
                    min(start_row + 3, main_ws.max_row + 1),
                ):
                    _set_value(main_ws, rr, col, None)

        total_val = round(
            sum(
                cv_map.get((d.normalize(), m), 0)
                for d in dates
                for m in members
            )
        )
        _set_value(main_ws, start_row, total_col, total_val)
        _set_value(main_ws, start_row + 1, total_col, 0)
        _set_value(main_ws, start_row + 2, total_col, -total_val)

    # 明細エリア
    # 高速テンプレート側に150媒体×4行の空枠・書式を事前作成済み。
    # ここでは行追加・全セルクリア・スタイルコピーを一切行わず、値だけ書く。
    detail_start = 21

    for idx, media in enumerate(media_list, start=0):
        r0 = detail_start + idx * 4

        sid = sid_map.get(media, "")
        media_type = media_type_map.get(media, "ポイントサイト")
        total_cv = round(
            sum(cv_map.get((d.normalize(), media), 0) for d in dates)
        )
        total_cost = round(
            sum(expected_cost_map.get((d.normalize(), media), 0) for d in dates)
        )
        overall_cpa = round(total_cost / total_cv) if total_cv else 0

        # 左側情報
        _set_value(main_ws, r0, 2, idx + 1)
        _set_value(main_ws, r0, 3, sid)
        _set_value(main_ws, r0, 4, media)
        _set_value(main_ws, r0, 7, total_cv)
        _set_value(main_ws, r0, 8, 0)
        _set_value(main_ws, r0, 10, 0)
        normal_rate = normal_rate_map.get(media, normal_rate_all)
        selected_rate = manual_rate_map.get(media, 0.0)
        _set_percent(main_ws, r0, 11, normal_rate)
        _set_percent(main_ws, r0, 12, selected_rate)
        _set_value(main_ws, r0, 17, total_cost)
        _set_value(main_ws, r0, 19, media_type)

        metrics = [
            "Daily Target (Initiative)",
            "Actural",
            "Promotion Detail",
            "GAP",
        ]
        for off, metric in enumerate(metrics):
            _set_value(main_ws, r0 + off, 24, metric)

        for i in range(date_slots):
            col = first_date_col + i
            if i < len(dates):
                dt = dates[i].normalize()
                cv = round(cv_map.get((dt, media), 0))
                unit_price = round(period_unit_map.get((dt, media), 0))

                _set_value(main_ws, r0, col, cv)
                _set_value(main_ws, r0 + 1, col, None)
                _set_value(main_ws, r0 + 2, col, unit_price)
                _set_value(main_ws, r0 + 3, col, -cv)
            else:
                for off in range(4):
                    _set_value(main_ws, r0 + off, col, None)

        _set_value(main_ws, r0, total_col, total_cv)
        _set_value(main_ws, r0 + 1, total_col, 0)
        _set_value(main_ws, r0 + 2, total_col, overall_cpa)
        _set_value(main_ws, r0 + 3, total_col, -total_cv)

    # =========================================================
    # 2) 件数 / 3) 発行 / 4) コスト計算
    # 3シートとも同じ列構造:
    # A SID / B 媒体名 / C 単価①最適 / D 単価②マジ得 /
    # E 定常承認率 / F 定常発行数 /
    # G マジ得承認率 / H マジ得発行数 /
    # I:AO 日次 / AP Total
    # =========================================================
    metric_specs = [
        ("件数(合算）", "count"),
        ("発行(合算）", "issue"),
        ("コスト計算用(合算）", "cost"),
    ]

    metric_first_date_col = 9   # I
    metric_total_col = 42       # AP
    metric_header_row = 3
    metric_data_start = 4

    for sheet_name, metric_kind in metric_specs:
        ws = wb_out[sheet_name]

        _set_date_slots(
            ws,
            metric_header_row,
            metric_first_date_col,
            date_slots,
            dates,
            metric_total_col,
        )

        # 固定列ヘッダーを毎回明示。
        headers = [
            "SID",
            "媒体名",
            "単価①\n最適プラン",
            "単価②\nマジ得",
            "定常承認率",
            "定常発行数",
            "マジ得承認率",
            "マジ得発行数",
        ]
        for c, header in enumerate(headers, start=1):
            _set_value(ws, metric_header_row, c, header)

        for idx, media in enumerate(media_list):
            r = metric_data_start + idx

            normal_rate = normal_rate_map.get(media, normal_rate_all)
            manual_rate = manual_rate_map.get(media, 0.0)

            opt_unit = manual_gross_unit_map.get(media, 0.0)
            magi_unit = magi_unit_map.get(media, magi_unit_all)
            if magi_unit <= 0:
                magi_unit = opt_unit

            _set_value(ws, r, 1, sid_map.get(media, ""))
            _set_value(ws, r, 2, media)

            # Cは今回の手動採用グロス単価。
            _set_value(ws, r, 3, round(opt_unit))
            _set_value(ws, r, 4, round(magi_unit))

            # Eは過去定常の参考値、Gは今回採用承認率を反映。
            _set_percent(ws, r, 5, normal_rate)
            _set_value(ws, r, 6, round(normal_issue_by_media.get(media, 0)))
            _set_percent(ws, r, 7, manual_rate)
            _set_value(ws, r, 8, round(magi_issue_by_media.get(media, 0)))

            row_total = 0.0

            for i in range(date_slots):
                c = metric_first_date_col + i

                if i >= len(dates):
                    _set_value(ws, r, c, None)
                    continue

                dt = dates[i].normalize()

                if metric_kind == "count":
                    value = cv_map.get((dt, media), 0)
                elif metric_kind == "issue":
                    value = issue_map.get((dt, media), 0)
                else:
                    value = expected_cost_map.get((dt, media), 0)

                value = round(value)
                _set_value(ws, r, c, value)
                row_total += value

            _set_value(ws, r, metric_total_col, round(row_total))

    # =========================================================
    # 5) 全体サマリ
    # =========================================================
    sws = wb_out["全体サマリ（定常期間サマリ）"]
    _set_value(
        sws,
        1,
        1,
        f"{pd.Timestamp(start_date).month}月度サマリ 件数・発行・コスト",
    )

    sum_forecast = 0
    sum_issue = 0
    sum_cost = 0

    for i in range(33):
        r = 3 + i

        if i < len(dates):
            dt = dates[i].normalize()
            forecast = round(total_cv_by_date.get(dt, 0))
            issue = round(total_issue_by_date.get(dt, 0))
            cost = round(total_cost_by_date.get(dt, 0))

            approval_rate = issue / forecast if forecast else 0
            issue_cpa = round(cost / issue) if issue else 0

            _set_value(sws, r, 1, dates[i].to_pydatetime())
            if not isinstance(sws.cell(r, 1), MergedCell):
                sws.cell(r, 1).number_format = "m/d"

            # A日付 B目標 CForecast DGAP E発行 F発行GAP G承認率 H発行コスト I発行CPA
            _set_value(sws, r, 2, forecast)
            _set_value(sws, r, 3, forecast)
            _set_value(sws, r, 4, 0)
            _set_value(sws, r, 5, issue)
            _set_value(sws, r, 6, 0)
            _set_percent(sws, r, 7, approval_rate)
            _set_value(sws, r, 8, cost)
            _set_value(sws, r, 9, issue_cpa)

            sum_forecast += forecast
            sum_issue += issue
            sum_cost += cost
        else:
            for c in range(1, 10):
                _set_value(sws, r, c, None)

    total_rate = sum_issue / sum_forecast if sum_forecast else 0
    total_cpa = round(sum_cost / sum_issue) if sum_issue else 0

    _set_value(sws, 36, 1, "Total")
    _set_value(sws, 36, 2, sum_forecast)
    _set_value(sws, 36, 3, sum_forecast)
    _set_value(sws, 36, 4, 0)
    _set_value(sws, 36, 5, sum_issue)
    _set_value(sws, 36, 6, 0)
    _set_percent(sws, 36, 7, total_rate)
    _set_value(sws, 36, 8, sum_cost)
    _set_value(sws, 36, 9, total_cpa)

    # =========================================================
    # 既存移管合算シート：不要列 E:W を最終出力時に削除
    # =========================================================
    # すべての値を書き込み終えた後に削除するため、
    # 既存の列番号ベースの出力処理には影響しない。
    main_ws.delete_cols(5, 19)  # E:W（19列）

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output.getvalue()


# -----------------------
# ✅ 提出用Excelダウンロード専用Fragment
# -----------------------
def _submission_download_body(
    opt_summary,
    history_df,
    cpn_master,
    manual_settings,
    start_date,
    end_date,
    selected_cpn,
    opt_mode,
    submission_filename,
):
    """
    提出用Excelはクリックされた時だけ生成する。
    data に callable を渡すため、画面描画時にはExcelを作らない。
    """

    def build_submission_excel():
        return create_submission_excel(
            opt_summary=opt_summary,
            history_df=history_df,
            cpn_master=cpn_master,
            manual_settings=manual_settings,
            start_date=start_date,
            end_date=end_date,
            selected_cpn=selected_cpn,
            opt_mode=opt_mode,
        )

    st.download_button(
        "📥 提出用Excelを生成してDL",
        data=build_submission_excel,
        file_name=submission_filename,
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        on_click="ignore",
        type="primary",
        width="stretch",
    )


# st.fragment が使えるStreamlitでは、
# このボタン操作だけを独立再実行する。
if hasattr(st, "fragment"):
    render_submission_download = st.fragment(
        _submission_download_body
    )
else:
    # 古いStreamlitでも起動自体は可能。
    render_submission_download = _submission_download_body


# -----------------------
# ✅ 補助関数
# -----------------------
def _truthy(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.lower().isin(
        {"1", "true", "yes", "y", "○", "〇", "あり", "有", "実施"}
    )


def _normalize_sid(value) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _figure_png_bytes(fig, dpi: int = 180) -> bytes:
    buffer = BytesIO()
    fig.savefig(buffer, format="png", dpi=dpi, bbox_inches="tight")
    buffer.seek(0)
    return buffer.getvalue()


def _get_japanese_font_properties():
    """Noto CJK等の日本語フォントをファイルから直接指定する。

    Streamlit CloudではMatplotlibのフォントキャッシュに新規インストール済み
    フォントが反映されない場合があるため、family名ではなくfnameを使う。
    """
    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJKjp-Regular.otf",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/fonts-japanese-gothic.ttf",
    ]
    candidates.extend(glob.glob("/usr/share/fonts/**/*NotoSansCJK*", recursive=True))
    candidates.extend(glob.glob("/usr/share/fonts/**/*NotoSansJP*", recursive=True))
    candidates.extend(glob.glob("/usr/share/fonts/**/*IPA*Gothic*", recursive=True))

    seen = set()
    for path in candidates:
        if path in seen:
            continue
        seen.add(path)
        try:
            if Path(path).is_file():
                prop = font_manager.FontProperties(fname=path)
                # 実際に名前を取得できることまで確認
                _ = prop.get_name()
                plt.rcParams["axes.unicode_minus"] = False
                return prop
        except Exception:
            continue
    return None


def _render_png_actions(png_bytes: bytes, file_name: str, download_key: str, copy_key: str):
    """PNG保存とクリップボードへの画像コピーを横並びで表示する。"""
    left, right = st.columns([1, 1])
    with left:
        st.download_button(
            "画像をPNG保存",
            data=png_bytes,
            file_name=file_name,
            mime="image/png",
            key=download_key,
            width="stretch",
        )
    with right:
        b64 = base64.b64encode(png_bytes).decode("ascii")
        button_id = f"copy_{copy_key}".replace("-", "_")
        html = f"""
        <div style="width:100%;">
          <button id="{button_id}" style="
            width:100%; height:38px; border:1px solid rgba(49,51,63,.2);
            border-radius:8px; background:white; cursor:pointer; font-size:14px;
          ">画像をコピー</button>
          <div id="{button_id}_msg" style="font-size:12px; margin-top:3px; min-height:16px;"></div>
        </div>
        <script>
        const btn = document.getElementById('{button_id}');
        const msg = document.getElementById('{button_id}_msg');
        btn.addEventListener('click', async () => {{
          try {{
            const response = await fetch('data:image/png;base64,{b64}');
            const blob = await response.blob();
            if (!navigator.clipboard || typeof ClipboardItem === 'undefined') {{
              throw new Error('clipboard_api_unavailable');
            }}
            await navigator.clipboard.write([new ClipboardItem({{'image/png': blob}})]);
            msg.textContent = 'コピーしました';
          }} catch (e) {{
            msg.textContent = 'ブラウザの権限でコピーできませんでした';
          }}
        }});
        </script>
        """
        components.html(html, height=62)


def _chart_month_label(value) -> str:
    """日本語フォントが無い場合にも文字化けしない月度ラベルへ変換する。"""
    text = str(value).strip()
    import re
    match = re.search(r"(\d{4})年\s*(\d{1,2})月度?", text)
    if match:
        return f"{match.group(1)}/{int(match.group(2)):02d}"
    return text


def render_history_analytics(
    history_df: pd.DataFrame,
    current_plan_df: pd.DataFrame | None = None,
    current_plan_magitoku_df: pd.DataFrame | None = None,
):
    st.subheader("📈 月度別 過去実績分析")
    st.caption(
        "月度はCPNマスタの『月度』を正として集計。"
        "発行数 = 成果承認フラグYの件数、発行CPA = コスト ÷ 発行数です。"
    )

    analysis_scope = st.radio(
        "集計対象",
        ["全期間", "マジ得期間のみ"],
        horizontal=True,
        key="history_analytics_scope",
    )

    analytics_df = history_df
    current_chart_plan = current_plan_df
    scope_suffix = "全期間"
    if analysis_scope == "マジ得期間のみ":
        if "CPN名" not in history_df.columns:
            st.warning("CPN名が付与されていないため、マジ得期間のみの集計ができません。")
            return
        analytics_df = history_df.loc[
            history_df["CPN名"].astype(str).str.strip().eq("マジ得")
        ].copy()
        current_chart_plan = current_plan_magitoku_df
        scope_suffix = "マジ得期間のみ"
        if analytics_df.empty:
            st.info("マジ得期間に該当する過去実績がありません。")
            return

    monthly = prepare_monthly_performance(analytics_df)
    if monthly.empty:
        st.info("月度が付与された過去実績がないため、月度別グラフを表示できません。")
        return

    preview = monthly[["月度", "発生件数", "発行数", "コスト", "発行CPA"]].copy()
    preview["発行CPA"] = preview["発行CPA"].map(
        lambda x: "" if pd.isna(x) else f"¥{x:,.0f}"
    )
    preview["コスト"] = preview["コスト"].map(lambda x: f"¥{x:,.0f}")

    tab_issue, tab_band, tab_table = st.tabs([
        "発行数・発行CPA",
        "単価帯",
        "集計表",
    ])

    with tab_issue:
        jp_font = _get_japanese_font_properties()
        x = range(len(monthly))
        labels_raw = monthly["月度"].astype(str).tolist()
        labels = labels_raw if jp_font else [_chart_month_label(v) for v in labels_raw]

        fig_count, ax_count = plt.subplots(figsize=(12, 5))
        ax_count.bar(x, monthly["発行数"])
        ax_count.set_title(
            f"月度別 発行数（{scope_suffix}）" if jp_font else f"Issued count by month ({scope_suffix})",
            fontproperties=jp_font,
        )
        ax_count.set_ylabel("発行数" if jp_font else "Issued count", fontproperties=jp_font)
        ax_count.set_xticks(list(x))
        ax_count.set_xticklabels(labels, rotation=45, ha="right", fontproperties=jp_font)
        ax_count.grid(axis="y", alpha=0.25)
        fig_count.tight_layout()
        st.pyplot(fig_count, width="stretch")
        count_png = _figure_png_bytes(fig_count)
        _render_png_actions(
            count_png,
            f"月度別_発行数_{scope_suffix}.png",
            f"download_monthly_issue_count_{analysis_scope}",
            f"copy_monthly_issue_count_{analysis_scope}",
        )
        plt.close(fig_count)

        fig_cpa, ax_cpa = plt.subplots(figsize=(12, 5))
        ax_cpa.plot(x, monthly["発行CPA"], marker="o")
        ax_cpa.set_title(
            f"月度別 発行CPA（{scope_suffix}）" if jp_font else f"Issued CPA by month ({scope_suffix})",
            fontproperties=jp_font,
        )
        ax_cpa.set_ylabel("発行CPA（円）" if jp_font else "Issued CPA (JPY)", fontproperties=jp_font)
        ax_cpa.set_xticks(list(x))
        ax_cpa.set_xticklabels(labels, rotation=45, ha="right", fontproperties=jp_font)
        ax_cpa.grid(axis="y", alpha=0.25)
        fig_cpa.tight_layout()
        st.pyplot(fig_cpa, width="stretch")
        cpa_png = _figure_png_bytes(fig_cpa)
        _render_png_actions(
            cpa_png,
            f"月度別_発行CPA_{scope_suffix}.png",
            f"download_monthly_issue_cpa_{analysis_scope}",
            f"copy_monthly_issue_cpa_{analysis_scope}",
        )
        plt.close(fig_cpa)

    with tab_band:
        control1, control2 = st.columns([1, 1])
        with control1:
            band_step = st.number_input(
                "単価帯の刻み幅（円）",
                min_value=500,
                max_value=50000,
                value=4000,
                step=500,
            )
        with control2:
            band_metric = st.selectbox(
                "積み上げる指標",
                ["発行数", "発生件数"],
                index=0,
            )

        # 1,000円刻み等で単価帯が大量発生してもクラッシュしないよう、
        # 表示系列数は最大40帯に制御。超過分は高単価側を「○円以上」に集約する。
        band_matrix = prepare_unit_price_band_matrix(
            analytics_df,
            band_step=int(band_step),
            value_metric=band_metric,
            max_bands=40,
        )

        # 今回プラン（手動設定後）を同じ単価帯グラフへ追加。
        if current_chart_plan is not None and not current_chart_plan.empty:
            cp = _normalize_manual_settings(current_chart_plan)
            cp["unit_price"] = pd.to_numeric(
                cp["今回プラン採用グロス単価"], errors="coerce"
            ).fillna(0.0)
            metric_col = "承認件数" if band_metric == "発行数" else "今回採用件数"
            cp[metric_col] = pd.to_numeric(cp[metric_col], errors="coerce").fillna(0.0)
            cp = cp[(cp["unit_price"] >= 0) & (cp[metric_col] > 0)].copy()
            if not cp.empty:
                cp["band_lower"] = (
                    (cp["unit_price"] // int(band_step)).astype(int) * int(band_step)
                )
                current = (
                    cp.groupby("band_lower", as_index=False)[metric_col]
                    .sum()
                )
                current_labels = {}
                for r in current.itertuples():
                    lower = int(r.band_lower)
                    upper = lower + int(band_step) - 1
                    current_labels[lower] = f"¥{lower:,}–¥{upper:,}"

                # 既存履歴側の帯と今回帯を和集合にする。
                if band_matrix.empty:
                    band_matrix = pd.DataFrame(index=["今回"])
                    band_matrix.index.name = "月度"
                all_cols = list(band_matrix.columns)
                for label in current_labels.values():
                    if label not in all_cols:
                        all_cols.append(label)
                band_matrix = band_matrix.reindex(columns=all_cols, fill_value=0.0)
                band_matrix.loc["今回"] = 0.0
                for _, r in current.iterrows():
                    band_matrix.loc["今回", current_labels[int(r["band_lower"])]] = float(r[metric_col])

        # 今回分を足した結果も最大40帯に制御し、細かい刻みでのメモリ急増を防ぐ。
        if not band_matrix.empty and len(band_matrix.columns) > 40:
            import re
            def _band_lower_from_label(label):
                nums = re.findall(r"[\d,]+", str(label))
                return int(nums[0].replace(",", "")) if nums else 0
            ordered_cols = sorted(band_matrix.columns, key=_band_lower_from_label)
            keep_cols = ordered_cols[:39]
            overflow_cols = ordered_cols[39:]
            overflow_lower = _band_lower_from_label(overflow_cols[0])
            overflow_label = f"¥{overflow_lower:,}以上"
            overflow_values = band_matrix[overflow_cols].sum(axis=1)
            band_matrix = band_matrix[keep_cols].copy()
            band_matrix[overflow_label] = overflow_values

        if band_matrix.empty:
            st.info("単価帯グラフを作成できる実績がありません。")
        else:
            jp_font = _get_japanese_font_properties()
            plot_matrix = band_matrix.copy()
            if not jp_font:
                plot_matrix.index = [_chart_month_label(v) for v in plot_matrix.index]
            fig_band, ax_band = plt.subplots(figsize=(12, 6))
            plot_matrix.plot(kind="bar", stacked=True, ax=ax_band, width=0.8)
            ax_band.set_title(
                f"月度別 単価帯構成（{scope_suffix} / {int(band_step):,}円刻み / {band_metric}）"
                if jp_font
                else f"Unit price bands by month ({int(band_step):,} JPY step)",
                fontproperties=jp_font,
            )
            ax_band.set_xlabel("月度" if jp_font else "Month", fontproperties=jp_font)
            ax_band.set_ylabel(band_metric if jp_font else ("Issued count" if band_metric == "発行数" else "Conversions"), fontproperties=jp_font)
            ax_band.tick_params(axis="x", rotation=45)
            if jp_font:
                for tick in ax_band.get_xticklabels():
                    tick.set_fontproperties(jp_font)
            # 凡例の縦長化でPNGが巨大化しないよう、系列数に応じて複数列化。
            legend_cols = max(1, min(4, (len(plot_matrix.columns) + 14) // 15))
            legend_prop = jp_font.copy() if jp_font is not None else None
            if legend_prop is not None:
                legend_prop.set_size(8)
            legend = ax_band.legend(
                title="単価帯" if jp_font else "Unit price band",
                bbox_to_anchor=(1.02, 1),
                loc="upper left",
                ncol=legend_cols,
                fontsize=8,
                prop=legend_prop,
            )
            if jp_font and legend is not None:
                legend.get_title().set_fontproperties(jp_font)
            ax_band.grid(axis="y", alpha=0.25)
            fig_band.tight_layout()
            st.pyplot(fig_band, width="stretch")
            # 積み上げグラフは系列数が多いため、PNG生成時のメモリ使用量も抑える。
            band_png = _figure_png_bytes(fig_band, dpi=140)
            _render_png_actions(
                band_png,
                f"月度別_単価帯_{int(band_step)}円刻み_{scope_suffix}.png",
                f"download_unit_price_band_{analysis_scope}",
                f"copy_unit_price_band_{analysis_scope}",
            )
            plt.close(fig_band)

    with tab_table:
        st.dataframe(preview, width="stretch", hide_index=True)


def _daily_pair_average(df: pd.DataFrame) -> pd.DataFrame:
    daily = (
        df.groupby(["date", "media", "商品ID"], as_index=False)
        .agg(cv=("cv", "sum"), cost=("cost", "sum"))
    )
    return (
        daily.groupby(["media", "商品ID"], as_index=False)
        .agg(base_cv=("cv", "mean"), cost=("cost", "mean"))
    )


def _calculate_normal_month_base(
    history_df: pd.DataFrame,
    selected_months: list[str],
) -> pd.DataFrame:
    """選択した月度の定常/通常実績から媒体×商品IDの日平均を返す。

    この処理はapp.py内に置き、logic.factorsの更新タイミング差による
    ImportErrorでアプリ全体が起動不能になるのを防ぐ。
    """
    columns = ["media", "商品ID", "base_cv", "cost"]
    if not selected_months:
        return pd.DataFrame(columns=columns)

    required = {"date", "media", "商品ID", "月度", "CPN名", "cv", "cost"}
    missing = required - set(history_df.columns)
    if missing:
        raise ValueError(
            "定常月度学習に必要な列がありません: " + ", ".join(sorted(missing))
        )

    work = history_df.copy()
    work["date"] = pd.to_datetime(work["date"], errors="coerce").dt.normalize()
    work["月度"] = work["月度"].astype("string").str.strip()
    work["CPN名"] = work["CPN名"].astype("string").str.strip()

    selected_month_set = {str(x).strip() for x in selected_months}
    selected = work[
        work["月度"].isin(selected_month_set)
        & work["CPN名"].isin(["通常", "定常"])
    ].copy()
    if selected.empty:
        return pd.DataFrame(columns=columns)

    total_days = selected["date"].dropna().nunique()
    if total_days <= 0:
        return pd.DataFrame(columns=columns)

    result = (
        selected.groupby(["media", "商品ID"], as_index=False)
        .agg(total_cv=("cv", "sum"), total_cost=("cost", "sum"))
    )
    result["base_cv"] = result["total_cv"] / float(total_days)
    result["cost"] = result["total_cost"] / float(total_days)
    return result[columns]


# -----------------------
# ✅ UI
# -----------------------
st.set_page_config(page_title="AFプランニングツール", layout="wide")
st.title("📊 件数予測＆プランニングツール")
st.caption(
    "実績CSVと最新のCPNマスタをアップロードしてください。"
    "CPNマスタ内の『CPNマスタ』『媒体名マスタ』シートを使用します。"
    "ファイルはGitHubには保存されません。"
)

col1, col2 = st.columns(2)

with col1:
    uploaded_file = st.file_uploader("① 実績CSV", type=["csv"])

with col2:
    uploaded_master = st.file_uploader(
        "② CPNマスタ",
        type=["xlsx", "xlsm"],
        help="『CPNマスタ』『媒体名マスタ』の2シートを含むExcelをアップロードしてください。",
    )

if uploaded_file and uploaded_master:
    st.sidebar.header("⚙️ 実績データ条件")
    exclude_compensation = st.sidebar.toggle(
        "補填を除外",
        value=True,
        help="ON: 実績データP列に文字列が入っている補填対象行を除外します。",
    )

    try:
        from data.loader import add_business_edge_flags
        from logic.factors import (
            calculate_dynamic_factor_tables,
            calculate_selected_cpn_base,
            get_cpn_reference_periods,
            enforce_premium_media_cost,
        )
        from config.constants import RECENT_NORMAL_DAYS

        # Streamlitはウィジェット操作のたびにスクリプト全体を再実行する。
        # 60MB級CSVをその都度pd.read_csvすると瞬間的にメモリを大きく消費するため、
        # ファイル内容をキーに整形済みDataFrameをキャッシュする。
        history_bytes = uploaded_file.getvalue()
        master_bytes = uploaded_master.getvalue()

        history_df = _load_history_cached(
            history_bytes,
            exclude_compensation=exclude_compensation,
        )
        history_df["date"] = pd.to_datetime(
            history_df["date"],
            errors="coerce",
        ).dt.normalize()

        # CPNマスタ/媒体名マスタも再実行のたびにopenpyxlで読み直さない。
        cpn_master, media_master = _load_master_cached(master_bytes)

        required_master_columns = {"日付", "CPN名", "月度"}
        missing_master = required_master_columns - set(cpn_master.columns)

        if missing_master:
            raise ValueError(
                f"CPNマスタに必要な列がありません: "
                f"{', '.join(sorted(missing_master))}"
            )

        cpn_master = cpn_master.copy()
        cpn_master["日付"] = pd.to_datetime(
            cpn_master["日付"],
            errors="coerce",
        ).dt.normalize()
        cpn_master["CPN名"] = (
            cpn_master["CPN名"]
            .astype("string")
            .str.strip()
        )
        cpn_master["月度"] = (
            cpn_master["月度"]
            .astype("string")
            .str.strip()
        )
        cpn_master = cpn_master.dropna(
            subset=["日付", "CPN名"]
        )
        cpn_master["月度"] = (
            cpn_master["月度"]
            .replace("", pd.NA)
            .fillna("未設定")
        )
        cpn_master = cpn_master.drop_duplicates(
            subset=["日付"],
            keep="last",
        )

        if cpn_master.empty:
            raise ValueError(
                "CPNマスタに有効な日付・CPN名がありません。"
            )

        # 任意列。未登録なら補正なし。
        cpn_master["line_oa_flag"] = (
            _truthy(cpn_master["LINE OA配信"])
            if "LINE OA配信" in cpn_master
            else 0
        )
        cpn_master["magitoku_after_flag"] = (
            _truthy(cpn_master["マジ得後"])
            if "マジ得後" in cpn_master
            else 0
        )

        # 媒体名マスタは上のキャッシュ読込で取得済み。

    except Exception as exc:
        st.error(f"ファイルの読み込みに失敗しました: {exc}")
        st.stop()

    history_df = history_df.merge(
        cpn_master[
            [
                "日付",
                "CPN名",
                "月度",
                "line_oa_flag",
                "magitoku_after_flag",
            ]
        ],
        left_on="date",
        right_on="日付",
        how="left",
    )

    history_df["CPN名"] = history_df["CPN名"].fillna("通常")
    history_df["line_oa_flag"] = (
        history_df["line_oa_flag"]
        .fillna(0)
        .astype(int)
    )
    history_df["magitoku_after_flag"] = (
        history_df["magitoku_after_flag"]
        .fillna(0)
        .astype(int)
    )
    history_df["月度"] = (
        history_df["月度"]
        .astype("string")
        .str.strip()
        .replace("", pd.NA)
        .fillna("未設定")
    )

    history_df = history_df.merge(
        media_master[["SID", "媒体名", "カテゴリ"]],
        on="SID",
        how="left",
    )
    history_df["raw_media"] = history_df["media"]
    mapped_media = (
        history_df["媒体名"]
        .astype("string")
        .str.strip()
        .replace("", pd.NA)
    )
    history_df["media"] = mapped_media.fillna(history_df["raw_media"])
    history_df["media_category"] = (
        history_df["カテゴリ"]
        .astype("string")
        .str.strip()
        .replace("", pd.NA)
        .fillna("未分類")
    )

    st.sidebar.header("商品ID選択")
    all_product_ids = sorted(
        [x for x in history_df["商品ID"].dropna().astype(str).unique() if str(x).strip() != ""]
    )
    product_id_one = next(
        (pid for pid in all_product_ids if str(pid).strip() == "1"),
        next(
            (pid for pid in all_product_ids if pd.to_numeric(pid, errors="coerce") == 1),
            None,
        ),
    )
    default_product_ids = [product_id_one] if product_id_one is not None else all_product_ids[:1]
    selected_product_ids = st.sidebar.multiselect(
        "商品ID",
        all_product_ids,
        default=default_product_ids,
    )
    if not selected_product_ids:
        st.stop()

    history_df = history_df[
        history_df["商品ID"].astype(str).isin(selected_product_ids)
    ].copy()

    st.sidebar.header("媒体選択")
    all_categories = sorted(history_df["media_category"].dropna().astype(str).unique())
    selected_categories = st.sidebar.multiselect(
        "カテゴリ",
        all_categories,
        default=all_categories,
    )
    if not selected_categories:
        st.stop()

    history_df = history_df[
        history_df["media_category"].isin(selected_categories)
    ].copy()

    all_media = sorted(history_df["media"].dropna().astype(str).unique())
    default_media = [
        m for m in all_media
        if "計測" not in m
    ]

    selected_media = st.sidebar.multiselect(
        "媒体",
        all_media,
        default=default_media,
    )

    if not selected_media:
        st.stop()

    history_df = history_df[
        history_df["media"].isin(selected_media)
    ].copy()

    # 月度別過去実績分析はUI上部へ表示する。
    # 今回プランは後段で確定するため、ここでは表示位置だけ確保し、
    # 手動設定取得後にこのプレースホルダーへ描画する。
    history_analytics_placeholder = st.empty()

    # 定常 / マジ得の承認率を画面でも確認。
    try:
        _period_preview = _calculate_period_media_metrics(history_df)
        approval_preview = pd.DataFrame(
            [
                {
                    "媒体": media,
                    "定常承認率": _period_preview["normal_rate"].get(
                        media,
                        _period_preview["normal_rate_all"],
                    ),
                    "マジ得承認率": _period_preview["magi_rate"].get(
                        media,
                        _period_preview["magi_rate_all"],
                    ),
                    "マジ得単価": _period_preview["magi_unit"].get(
                        media,
                        _period_preview["magi_unit_all"],
                    ),
                }
                for media in selected_media
            ]
        )

        approval_preview["定常承認率"] = approval_preview["定常承認率"].map(
            lambda x: f"{x:.1%}"
        )
        approval_preview["マジ得承認率"] = approval_preview["マジ得承認率"].map(
            lambda x: f"{x:.1%}"
        )
        approval_preview["マジ得単価"] = approval_preview["マジ得単価"].map(
            lambda x: f"¥{x:,.0f}"
        )

        with st.expander("✅ 定常・マジ得の過去実績指標"):
            st.caption(
                "承認率 = 成果承認フラグYの件数 ÷ 全件数 / "
                "マジ得単価 = マジ得期間cost ÷ CV"
            )
            st.dataframe(
                approval_preview,
                width="stretch",
                hide_index=True,
            )

    except ValueError as approval_exc:
        st.warning(
            "定常・マジ得指標を算出できません。"
            f"{approval_exc}"
        )

    st.sidebar.header("📊 対象期間・施策")

    cpn_list = sorted(
        cpn_master["CPN名"]
        .dropna()
        .astype(str)
        .unique()
    )
    if not cpn_list:
        st.error("CPNマスタに選択可能な施策がありません。")
        st.stop()

    normal_labels = {"通常", "定常"}
    default_primary = next((x for x in ["定常", "通常"] if x in cpn_list), cpn_list[0])
    today = datetime.date.today()

    selected_cpn_1 = st.sidebar.selectbox(
        "施策①",
        cpn_list,
        index=cpn_list.index(default_primary),
        key="planning_cpn_1",
    )
    start_date_1 = st.sidebar.date_input(
        "施策① 開始",
        today,
        key="planning_start_1",
    )
    end_date_1 = st.sidebar.date_input(
        "施策① 終了",
        today + datetime.timedelta(days=7),
        key="planning_end_1",
    )

    use_second_period = st.sidebar.toggle(
        "施策②を追加",
        value=False,
        key="planning_use_second",
    )

    selected_cpn_2 = None
    start_date_2 = None
    end_date_2 = None
    if use_second_period:
        default_second = "マジ得" if "マジ得" in cpn_list else cpn_list[0]
        selected_cpn_2 = st.sidebar.selectbox(
            "施策②",
            cpn_list,
            index=cpn_list.index(default_second),
            key="planning_cpn_2",
        )
        start_date_2 = st.sidebar.date_input(
            "施策② 開始",
            end_date_1 + datetime.timedelta(days=1),
            key="planning_start_2",
        )
        end_date_2 = st.sidebar.date_input(
            "施策② 終了",
            end_date_1 + datetime.timedelta(days=7),
            key="planning_end_2",
        )

    planning_segments = [
        {"label": "施策①", "cpn": selected_cpn_1, "start": start_date_1, "end": end_date_1}
    ]
    if use_second_period:
        planning_segments.append(
            {"label": "施策②", "cpn": selected_cpn_2, "start": start_date_2, "end": end_date_2}
        )

    for seg in planning_segments:
        if seg["start"] > seg["end"]:
            st.error(f"{seg['label']}の開始日は終了日以前にしてください。")
            st.stop()

    if use_second_period:
        r1 = pd.date_range(start_date_1, end_date_1)
        r2 = pd.date_range(start_date_2, end_date_2)
        if len(r1.intersection(r2)) > 0:
            st.error("施策①と施策②の期間が重複しています。期間が重ならないように設定してください。")
            st.stop()

    # 定常の基礎値は、選択月度の定常実績の日平均から算出する。
    needs_normal_learning = any(seg["cpn"] in normal_labels for seg in planning_segments)
    selected_learning_months = []
    if needs_normal_learning:
        normal_month_source = history_df[
            history_df["CPN名"].isin(normal_labels)
            & history_df["月度"].astype(str).ne("未設定")
        ][["date", "月度"]].copy()
        normal_month_order = (
            normal_month_source.groupby("月度", as_index=False)
            .agg(last_date=("date", "max"))
            .sort_values(["last_date", "月度"], kind="stable")
        )
        month_options = normal_month_order["月度"].astype(str).tolist()
        default_months = month_options[-6:] if len(month_options) > 6 else month_options

        selected_learning_months = st.sidebar.multiselect(
            "定常 学習月度（複数選択）",
            options=month_options,
            default=default_months,
            help="デフォルトは実績で読めた直近6月度。選択月度内の定常CV合計÷定常日数を基礎日平均にします。",
        )
        if not selected_learning_months:
            st.warning("定常を使う場合は学習月度を1つ以上選択してください。")
            st.stop()

    segment_bases = []
    reference_descriptions = []

    for seg_idx, seg in enumerate(planning_segments, start=1):
        selected_cpn = seg["cpn"]
        if selected_cpn in normal_labels:
            base_pair_seg = _calculate_normal_month_base(
                history_df,
                selected_learning_months,
            )
            if base_pair_seg.empty:
                st.error("選択した月度に定常実績がありません。")
                st.stop()
            reference_key_seg = ("normal_months", tuple(selected_learning_months))
            reference_descriptions.append(
                f"{seg['label']} {selected_cpn}: 定常学習月度 " + ", ".join(selected_learning_months)
            )
        else:
            available_periods = get_cpn_reference_periods(history_df, selected_cpn)
            if not available_periods:
                st.error(f"実績内に『{selected_cpn}』のキャンペーン期間がありません。")
                st.stop()

            period_options = {
                f"{p_start.strftime('%Y/%m/%d')} ～ {p_end.strftime('%Y/%m/%d')}": (p_start, p_end)
                for p_start, p_end in available_periods
            }
            selected_period_labels = st.sidebar.multiselect(
                f"{seg['label']} {selected_cpn} 参照期間",
                options=list(period_options.keys()),
                default=list(period_options.keys()),
                key=f"cpn_reference_periods_{seg_idx}",
            )
            if not selected_period_labels:
                st.warning(f"{seg['label']}の参照期間を1つ以上選択してください。")
                st.stop()
            selected_periods = [period_options[label] for label in selected_period_labels]
            base_pair_seg = calculate_selected_cpn_base(
                history_df,
                selected_cpn,
                selected_periods,
            )
            if base_pair_seg.empty:
                st.error(f"{seg['label']}の選択参照期間に対象媒体の実績がありません。")
                st.stop()
            reference_key_seg = (selected_cpn, tuple(selected_period_labels))
            reference_descriptions.append(
                f"{seg['label']} {selected_cpn}: " + " / ".join(selected_period_labels)
            )

        base_pair_seg = base_pair_seg[base_pair_seg["media"].isin(selected_media)].copy()
        if base_pair_seg.empty:
            st.error(f"{seg['label']}の学習実績に対象媒体がありません。")
            st.stop()
        segment_bases.append(
            {**seg, "base_pair": base_pair_seg, "reference_key": reference_key_seg}
        )

    st.subheader("📈 予測ベース")
    st.caption(
        "定常は選択月度の定常日平均、その他施策は選択した過去CPN期間の日平均を使用します。"
    )
    for desc in reference_descriptions:
        st.write(f"・{desc}")

    base_preview_rows = []
    for seg in segment_bases:
        tmp = (
            seg["base_pair"].groupby("media", as_index=False)
            .agg(base_cv=("base_cv", "sum"), cost=("cost", "sum"))
        )
        tmp["施策"] = seg["label"] + "：" + seg["cpn"]
        base_preview_rows.append(tmp)
    base_preview = pd.concat(base_preview_rows, ignore_index=True)
    base_preview = base_preview.rename(
        columns={"media": "媒体", "base_cv": "基礎CV/日", "cost": "基礎COST/日"}
    )
    base_preview["基礎CV/日"] = base_preview["基礎CV/日"].round(2)
    base_preview["基礎COST/日"] = base_preview["基礎COST/日"].round(0)
    st.dataframe(
        base_preview[["施策", "媒体", "基礎CV/日", "基礎COST/日"]],
        width="stretch",
        hide_index=True,
    )

    # 既存関数との互換用。提出用の代表施策は施策①とするが、
    # 日別施策はfuture_df側へ保持して複合期間を正しく予測する。
    selected_cpn = selected_cpn_1
    start_date = min(seg["start"] for seg in planning_segments)
    end_date = max(seg["end"] for seg in planning_segments)
    base_pair = pd.concat([seg["base_pair"] for seg in segment_bases], ignore_index=True)

    st.sidebar.header("🎯 最適化ロジック")

    opt_mode = st.sidebar.radio(
        "最適基準",
        ["CPA最小", "CV最大"],
        index=0,
    )

    # ---------------------------------------------------------
    # 変動係数は同じ入力条件なら再計算しない。
    # Excel生成ボタン等によるStreamlit再実行でも再利用する。
    # ---------------------------------------------------------
    factor_cache_key = (
        tuple(selected_media),
        tuple(selected_product_ids),
        bool(exclude_compensation),
        len(history_df),
        str(history_df["date"].min()),
        str(history_df["date"].max()),
        float(pd.to_numeric(history_df["cv"], errors="coerce").fillna(0).sum()),
        float(pd.to_numeric(history_df["cost"], errors="coerce").fillna(0).sum()),
    )

    if st.session_state.get("_factor_cache_key") == factor_cache_key:
        factor_tables = st.session_state["_factor_tables"]
    else:
        factor_tables = calculate_dynamic_factor_tables(
            history_df
        )
        st.session_state["_factor_cache_key"] = factor_cache_key
        st.session_state["_factor_tables"] = factor_tables

    st.subheader("📐 実績から算出した変動係数")

    tab1, tab2, tab3, tab4 = st.tabs(
        ["曜日", "月初・月末", "需要期", "LINE OA"]
    )

    with tab1:
        st.dataframe(
            factor_tables["weekday"].round(3),
            width="stretch",
            hide_index=True,
        )

    with tab2:
        st.dataframe(
            factor_tables["month_edge"].round(3),
            width="stretch",
            hide_index=True,
        )

    with tab3:
        st.dataframe(
            factor_tables["season"].round(3),
            width="stretch",
            hide_index=True,
        )

    with tab4:
        if factor_tables["line_oa"].empty:
            st.info(
                "過去のLINE OA配信実績がないため、"
                "LINE OA係数は1.0です。"
            )
        else:
            st.dataframe(
                factor_tables["line_oa"].round(3),
                width="stretch",
                hide_index=True,
            )

    # ---------------------------------------------------------
    # 予測 → 松竹梅 → 最適化 は一度計算したら session_state に保持。
    # Streamlitのボタン押下ではスクリプト全体が再実行されるが、
    # 入力条件が同じならここでは再計算しない。
    # ---------------------------------------------------------
    reference_key = tuple(
        (
            seg["label"],
            seg["cpn"],
            str(seg["start"]),
            str(seg["end"]),
            seg["reference_key"],
        )
        for seg in segment_bases
    )

    calc_key = (
        tuple(selected_media),
        tuple(selected_product_ids),
        bool(exclude_compensation),
        reference_key,
        opt_mode,
        len(base_pair),
        round(float(pd.to_numeric(base_pair["base_cv"], errors="coerce").fillna(0).sum()), 6),
        round(float(pd.to_numeric(base_pair["cost"], errors="coerce").fillna(0).sum()), 2),
    )

    cached_calc = st.session_state.get("_planning_calc")

    if (
        cached_calc is not None
        and st.session_state.get("_planning_calc_key") == calc_key
    ):
        forecast_df = cached_calc["forecast_df"]
        sim_df = cached_calc["sim_df"]
        sim_summary = cached_calc["sim_summary"]
        opt_df = cached_calc["opt_df"]
        opt_summary = cached_calc["opt_summary"]

        sim_report_table = cached_calc.get("sim_report_table")
        opt_report_table = cached_calc.get("opt_report_table")

        if sim_report_table is None:
            sim_report_table = create_report_table(sim_summary)
            cached_calc["sim_report_table"] = sim_report_table

        if opt_report_table is None:
            opt_report_table = create_report_table(opt_summary)
            cached_calc["opt_report_table"] = opt_report_table

    else:
        future_parts = []
        for seg in segment_bases:
            future_dates = pd.date_range(start=seg["start"], end=seg["end"])
            part = pd.DataFrame({"date": future_dates}).merge(
                seg["base_pair"],
                how="cross",
            )
            part["weekday"] = part["date"].dt.day_name()
            part = add_business_edge_flags(part)

            # CPNマスタからは月度・特殊日フラグだけ取得。
            # 施策名そのものはUI選択を正とする。
            part = part.merge(
                cpn_master[
                    [
                        "日付",
                        "月度",
                        "line_oa_flag",
                        "magitoku_after_flag",
                    ]
                ],
                left_on="date",
                right_on="日付",
                how="left",
            )
            part["CPN名"] = seg["cpn"]
            part["planning_segment"] = seg["label"]
            part["月度"] = (
                part["月度"]
                .astype("string")
                .str.strip()
                .replace("", pd.NA)
                .fillna(pd.Series(part["date"].dt.strftime("%Y年%m月度"), index=part.index))
            )
            part["line_oa_flag"] = part["line_oa_flag"].fillna(0).astype(int)
            part["magitoku_after_flag"] = part["magitoku_after_flag"].fillna(0).astype(int)
            part["cpn_factor"] = 1.0
            future_parts.append(part)

        future_df = pd.concat(future_parts, ignore_index=True)
        forecast_df = forecast_cv(future_df, factor_tables)

        forecast_df = enforce_premium_media_cost(
            forecast_df
        )

        # simulate_plan用に日付表示形式を変換
        forecast_df = forecast_df.copy()
        forecast_df["date"] = format_date(
            forecast_df
        )

        sim_df = simulate_plan(
            forecast_df
        )

        sim_group_cols = ["date", "media", "plan"]
        if "CPN名" in sim_df.columns:
            sim_group_cols.append("CPN名")
        sim_summary = (
            sim_df.groupby(
                sim_group_cols,
                as_index=False,
            )
            .agg(
                cv=("cv", "sum"),
                cost=("cost", "sum"),
            )
        )

        sim_summary["cpa"] = (
            (sim_summary["cost"] / sim_summary["cv"])
            .replace(
                [float("inf"), float("-inf")],
                0,
            )
            .fillna(0)
        )

        sim_summary["date"] = format_date(
            sim_summary
        )

        opt_df = optimize_budget(
            sim_df,
            opt_mode,
        )

        opt_group_cols = ["date", "media", "plan"]
        if "CPN名" in opt_df.columns:
            opt_group_cols.append("CPN名")
        opt_summary = (
            opt_df.groupby(
                opt_group_cols,
                as_index=False,
            )
            .agg(
                cv=("cv", "sum"),
                cost=("cost", "sum"),
            )
        )

        opt_summary["cpa"] = (
            (opt_summary["cost"] / opt_summary["cv"])
            .replace(
                [float("inf"), float("-inf")],
                0,
            )
            .fillna(0)
        )

        opt_summary["date"] = format_date(
            opt_summary
        )

        st.session_state["_planning_calc_key"] = calc_key
        # 表示用テーブルもここで1回だけ作って保存する。
        # Excelボタン操作や他ウィジェット操作で同じ表を作り直さない。
        sim_report_table = create_report_table(sim_summary)
        opt_report_table = create_report_table(opt_summary)

        st.session_state["_planning_calc"] = {
            "forecast_df": forecast_df,
            "sim_df": sim_df,
            "sim_summary": sim_summary,
            "opt_df": opt_df,
            "opt_summary": opt_summary,
            "sim_report_table": sim_report_table,
            "opt_report_table": opt_report_table,
        }

        # 条件が変わって再計算した場合、古い提出Excelは破棄
        st.session_state.pop("submission_excel", None)

    # 係数確認用の明細
    with st.expander("予測係数の確認"):
        factor_cols = [
            "date",
            "media",
            "商品ID",
            "base_cv",
            "cpn_factor",
            "weekday_factor",
            "season_factor",
            "month_edge_factor",
            "after_factor",
            "line_factor",
            "forecast_cv",
            "cost",
        ]

        available_factor_cols = [
            c for c in factor_cols
            if c in forecast_df.columns
        ]

        st.dataframe(
            forecast_df[available_factor_cols],
            width="stretch",
            hide_index=True,
        )

    st.subheader("📊 松竹梅")
    st.dataframe(
        sim_report_table,
        width="stretch",
    )

    st.subheader("🚀 最適プラン")
    st.dataframe(
        opt_report_table,
        width="stretch",
    )

    st.subheader("✍️ 手動設定")
    st.caption(
        "初期値は上の最適プランと過去実績から自動設定。"
        "グロス単価・承認率・採用件数・費用を直接編集できます。"
        "承認件数と発行CPAは入力内容から自動計算します。"
    )

    render_manual_settings(
        opt_summary=opt_summary,
        history_df=history_df,
        selected_cpn=selected_cpn,
        calc_key=calc_key,
    )

    # 過去実績分析には、手動設定後の今回プランを「今回」として単価帯グラフへ追加する。
    current_manual_for_chart = _normalize_manual_settings(
        st.session_state.get(
            "_manual_settings",
            _build_manual_settings_defaults(
                opt_summary=opt_summary,
                history_df=history_df,
                selected_cpn=selected_cpn,
            ),
        )
    )

    # 「マジ得期間のみ」では、今回プランもマジ得指定期間分だけを表示する。
    # 手動設定値は最終版として維持し、媒体ごとの採用件数だけ
    # 今回予測に占めるマジ得CV比率で按分する。
    current_manual_magitoku_for_chart = pd.DataFrame()
    if (
        not current_manual_for_chart.empty
        and "CPN名" in opt_summary.columns
        and (opt_summary["CPN名"].astype(str).str.strip() == "マジ得").any()
    ):
        mix = opt_summary.copy()
        mix["media"] = mix["media"].astype(str)
        mix["cv"] = pd.to_numeric(mix["cv"], errors="coerce").fillna(0.0)
        mix["CPN名"] = mix["CPN名"].astype(str).str.strip()
        total_cv = mix.groupby("media")["cv"].sum()
        magi_cv = mix.loc[mix["CPN名"].eq("マジ得")].groupby("media")["cv"].sum()
        magi_ratio = (magi_cv / total_cv.replace(0, pd.NA)).fillna(0.0).clip(0, 1)

        current_manual_magitoku_for_chart = current_manual_for_chart.copy()
        current_manual_magitoku_for_chart["_magi_ratio"] = (
            current_manual_magitoku_for_chart["媒体名"].astype(str).map(magi_ratio).fillna(0.0)
        )
        current_manual_magitoku_for_chart["今回採用件数"] = (
            pd.to_numeric(
                current_manual_magitoku_for_chart["今回採用件数"], errors="coerce"
            ).fillna(0.0)
            * current_manual_magitoku_for_chart["_magi_ratio"]
        )
        current_manual_magitoku_for_chart = current_manual_magitoku_for_chart.drop(
            columns=["_magi_ratio"]
        )
        current_manual_magitoku_for_chart = _normalize_manual_settings(
            current_manual_magitoku_for_chart
        )
        current_manual_magitoku_for_chart = current_manual_magitoku_for_chart.loc[
            current_manual_magitoku_for_chart["今回採用件数"] > 0
        ].copy()

    try:
        with history_analytics_placeholder.container():
            render_history_analytics(
                history_df,
                current_plan_df=current_manual_for_chart,
                current_plan_magitoku_df=current_manual_magitoku_for_chart,
            )
    except ValueError as analytics_exc:
        with history_analytics_placeholder.container():
            st.warning(f"月度別実績を集計できません。{analytics_exc}")

    submission_filename = (
        f"【提出用】楽天カード"
        f"{pd.Timestamp(start_date).year}年"
        f"{pd.Timestamp(start_date).month}月"
        f"プランニング.xlsx"
    )

    # ---------------------------------------------------------
    # 提出用Excel操作は独立Fragment。
    # ボタンを押しても予測・松竹梅・最適化は再実行しない。
    # Excel自体もクリックされるまで生成しない。
    # ---------------------------------------------------------
    manual_settings_for_export = _normalize_manual_settings(
        st.session_state.get(
            "_manual_settings",
            _build_manual_settings_defaults(
                opt_summary=opt_summary,
                history_df=history_df,
                selected_cpn=selected_cpn,
            ),
        )
    )

    render_submission_download(
        opt_summary=opt_summary,
        history_df=history_df,
        cpn_master=cpn_master,
        manual_settings=manual_settings_for_export,
        start_date=start_date,
        end_date=end_date,
        selected_cpn=selected_cpn,
        opt_mode=opt_mode,
        submission_filename=submission_filename,
    )
